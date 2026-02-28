import logging
from dataclasses import dataclass
from datetime import datetime, timezone
from io import BytesIO
from typing import List, Optional

from sqlalchemy.orm import Session

from app.config import MAX_ATTENDANCE_BREAKS
from app.constants import InputType
from app.core.exceptions import ConflictError, ForbiddenError, NotFoundError
from app.core.utils import parse_hhmm_to_utc, seconds_to_hm
from app.crud import attendance as crud_att
from app.crud import attendance_break as crud_break
from app.crud import attendance_preset as crud_preset
from app.crud import user as crud_user
from app.models.attendance import Attendance
from app.schemas.attendance import AttendanceCreate, AttendanceUpdate

logger = logging.getLogger("app.services.attendance")

MAX_BREAKS = MAX_ATTENDANCE_BREAKS
MIN_DURATION_SECONDS = 60  # 1 minute minimum for attendance/break duration


@dataclass
class AttendanceStatusResult:
    is_clocked_in: bool
    current_attendance: Optional[Attendance] = None


def _validate_min_duration(start: datetime, end: datetime, label: str) -> None:
    """Raise ConflictError if the duration between start and end is less than 1 minute."""
    if (end - start).total_seconds() < MIN_DURATION_SECONDS:
        raise ConflictError(f"{label}は1分以上必要です")


def clock_in(db: Session, user_id: int, note: Optional[str] = None) -> Attendance:
    current = crud_att.get_current_attendance(db, user_id)
    if current:
        logger.warning("Clock-in rejected: user_id=%d already clocked in", user_id)
        raise ConflictError("Already clocked in")

    # 1日1回制限: 同日に既に出退勤記録がある場合は拒否
    # CRUDが datetime.now(timezone.utc).date() で日付を設定するため、ここもUTCに統一
    today = datetime.now(timezone.utc).date()
    existing = crud_att.get_attendance_by_date(db, user_id, today)
    if existing:
        logger.warning("Clock-in rejected: user_id=%d already has attendance for today", user_id)
        raise ConflictError("Already clocked in today")

    result = crud_att.clock_in(db, user_id, note)
    logger.info("Clock-in: user_id=%d, attendance_id=%d", user_id, result.id)
    return result


def clock_out(db: Session, user_id: int, note: Optional[str] = None) -> Attendance:
    current = crud_att.get_current_attendance(db, user_id)
    if not current:
        logger.warning("Clock-out rejected: user_id=%d not clocked in", user_id)
        raise ConflictError("Not clocked in")
    now = datetime.now(timezone.utc)
    _validate_min_duration(current.clock_in, now, "勤務時間")
    result = crud_att.clock_out(db, current, note)
    logger.info("Clock-out: user_id=%d, attendance_id=%d", user_id, result.id)
    return result


def get_status(db: Session, user_id: int) -> AttendanceStatusResult:
    current = crud_att.get_current_attendance(db, user_id)
    if current:
        return AttendanceStatusResult(is_clocked_in=True, current_attendance=current)
    return AttendanceStatusResult(is_clocked_in=False, current_attendance=None)


def list_attendances(
    db: Session, user_id: int, year: Optional[int] = None, month: Optional[int] = None
) -> List[Attendance]:
    logger.info("Listing attendances for user_id=%d (year=%s, month=%s)", user_id, year, month)
    return crud_att.get_attendances(db, user_id, year=year, month=month)


def get_attendance(db: Session, attendance_id: int, user_id: int) -> Attendance:
    att = crud_att.get_attendance(db, attendance_id)
    if not att or att.user_id != user_id:
        logger.warning("Attendance not found: id=%d", attendance_id)
        raise NotFoundError("Attendance not found")
    return att


def create_attendance(db: Session, user_id: int, data: AttendanceCreate) -> Attendance:
    existing = crud_att.get_attendance_by_date(db, user_id, data.date)
    if existing:
        raise ConflictError("Attendance already exists for this date")

    clock_in_dt = parse_hhmm_to_utc(data.date, data.clock_in)
    clock_out_dt = parse_hhmm_to_utc(data.date, data.clock_out) if data.clock_out else None

    if clock_out_dt:
        _validate_min_duration(clock_in_dt, clock_out_dt, "勤務時間")

    # Validate break durations before creating
    if data.breaks:
        for brk in data.breaks[:MAX_BREAKS]:
            if brk.end:
                _validate_min_duration(
                    parse_hhmm_to_utc(data.date, brk.start),
                    parse_hhmm_to_utc(data.date, brk.end),
                    "休憩時間",
                )

    result = crud_att.create_attendance(
        db,
        user_id=user_id,
        target_date=data.date,
        clock_in=clock_in_dt,
        clock_out=clock_out_dt,
        note=data.note,
    )

    # Create breaks if provided
    if data.breaks:
        for brk in data.breaks[:MAX_BREAKS]:
            break_start_dt = parse_hhmm_to_utc(data.date, brk.start)
            brk_obj = crud_break.create_break(db, result.id, break_start_dt)
            if brk.end:
                break_end_dt = parse_hhmm_to_utc(data.date, brk.end)
                crud_break.end_break(db, brk_obj, break_end_dt)
        db.refresh(result)

    logger.info("Manual attendance created: user_id=%d, date=%s, id=%d", user_id, data.date, result.id)
    return result


def _check_admin_lock(att: Attendance) -> None:
    """Raise ForbiddenError if the record was entered by an admin."""
    if att.input_type == InputType.ADMIN:
        raise ForbiddenError("管理者入力のレコードは変更できません")


def update_attendance(db: Session, attendance_id: int, user_id: int, data: AttendanceUpdate) -> Attendance:
    att = crud_att.get_attendance(db, attendance_id)
    if not att or att.user_id != user_id:
        raise NotFoundError("Attendance not found")
    _check_admin_lock(att)

    update_data = {}
    new_clock_in = att.clock_in
    new_clock_out = att.clock_out
    if data.clock_in is not None:
        new_clock_in = parse_hhmm_to_utc(att.date, data.clock_in)
        update_data["clock_in"] = new_clock_in
    if data.clock_out is not None:
        new_clock_out = parse_hhmm_to_utc(att.date, data.clock_out)
        update_data["clock_out"] = new_clock_out
    if data.note is not None:
        update_data["note"] = data.note

    # Validate minimum duration for the resulting clock_in/clock_out
    if new_clock_in and new_clock_out:
        _validate_min_duration(new_clock_in, new_clock_out, "勤務時間")

    # Validate break durations before applying
    if data.breaks is not None:
        for brk in data.breaks[:MAX_BREAKS]:
            if brk.end:
                _validate_min_duration(
                    parse_hhmm_to_utc(att.date, brk.start),
                    parse_hhmm_to_utc(att.date, brk.end),
                    "休憩時間",
                )

    result = crud_att.update_attendance(db, att, update_data)

    if data.breaks is not None:
        crud_break.delete_breaks_by_attendance_id(db, result.id)
        for brk in data.breaks[:MAX_BREAKS]:
            break_start_dt = parse_hhmm_to_utc(result.date, brk.start)
            brk_obj = crud_break.create_break(db, result.id, break_start_dt)
            if brk.end:
                break_end_dt = parse_hhmm_to_utc(result.date, brk.end)
                crud_break.end_break(db, brk_obj, break_end_dt)
        db.refresh(result)

    logger.info("Attendance updated: id=%d, user_id=%d", attendance_id, user_id)
    return result


def delete_attendance(db: Session, attendance_id: int, user_id: int) -> None:
    att = crud_att.get_attendance(db, attendance_id)
    if not att or att.user_id != user_id:
        raise NotFoundError("Attendance not found")
    _check_admin_lock(att)
    crud_att.delete_attendance(db, att)
    logger.info("Attendance deleted: id=%d, user_id=%d", attendance_id, user_id)


def start_break(db: Session, attendance_id: int, user_id: int) -> Attendance:
    """Start a break for the given attendance record."""
    att = crud_att.get_attendance(db, attendance_id)
    if not att or att.user_id != user_id:
        raise NotFoundError("Attendance not found")
    _check_admin_lock(att)

    if att.clock_out is not None:
        raise ConflictError("Already clocked out")

    active = crud_break.get_active_break(db, attendance_id)
    if active:
        raise ConflictError("Break already active")

    count = crud_break.count_breaks(db, attendance_id)
    if count >= MAX_BREAKS:
        raise ConflictError(f"Maximum {MAX_BREAKS} breaks allowed")

    now = datetime.now(timezone.utc)
    crud_break.create_break(db, attendance_id, now)
    db.refresh(att)
    logger.info("Break started: attendance_id=%d, user_id=%d", attendance_id, user_id)
    return att


def end_break(db: Session, attendance_id: int, user_id: int) -> Attendance:
    """End the active break for the given attendance record."""
    att = crud_att.get_attendance(db, attendance_id)
    if not att or att.user_id != user_id:
        raise NotFoundError("Attendance not found")
    _check_admin_lock(att)

    active = crud_break.get_active_break(db, attendance_id)
    if not active:
        raise ConflictError("No active break")

    now = datetime.now(timezone.utc)
    _validate_min_duration(active.break_start, now, "休憩時間")
    crud_break.end_break(db, active, now)
    db.refresh(att)
    logger.info("Break ended: attendance_id=%d, user_id=%d", attendance_id, user_id)
    return att


def get_user_preset_id(db: Session, user_id: int) -> Optional[int]:
    user = crud_user.get_user(db, user_id)
    return user.default_preset_id if user else None


def set_user_preset_id(db: Session, user_id: int, preset_id: int) -> None:
    preset = crud_preset.get_preset(db, preset_id)
    if not preset:
        raise NotFoundError("Preset not found")
    user = crud_user.get_user(db, user_id)
    if not user:
        raise NotFoundError("User not found")
    user.default_preset_id = preset_id
    db.commit()
    logger.info("User %d default preset set to %d", user_id, preset_id)


def default_set(db: Session, user_id: int) -> Attendance:
    """Set today's attendance from user's default preset.

    Clock_in and clock_out are set from the preset.
    A single break is created from the preset's break_start/break_end.
    If a record already exists for today, it is updated.
    """
    user = crud_user.get_user(db, user_id)
    preset_id = user.default_preset_id if user and user.default_preset_id else 1
    preset = crud_preset.get_preset(db, preset_id)
    if not preset:
        raise NotFoundError("Preset not found")

    # Use UTC date to match CRUD layer's datetime.now(timezone.utc).date()
    today = datetime.now(timezone.utc).date()
    clock_in_dt = parse_hhmm_to_utc(today, preset.clock_in)
    clock_out_dt = parse_hhmm_to_utc(today, preset.clock_out)
    break_start_dt = parse_hhmm_to_utc(today, preset.break_start) if preset.break_start else None
    break_end_dt = parse_hhmm_to_utc(today, preset.break_end) if preset.break_end else None

    existing = crud_att.get_attendance_by_date(db, user_id, today)
    if existing:
        _check_admin_lock(existing)
        update_data = {
            "clock_in": clock_in_dt,
            "clock_out": clock_out_dt,
        }
        result = crud_att.update_attendance(db, existing, update_data)

        # Replace existing breaks with preset break
        crud_break.delete_breaks_by_attendance_id(db, result.id)
        if break_start_dt:
            brk = crud_break.create_break(db, result.id, break_start_dt)
            if break_end_dt:
                crud_break.end_break(db, brk, break_end_dt)

        logger.info("Default-set: updated attendance id=%d, preset=%s, user_id=%d", result.id, preset.name, user_id)
    else:
        result = crud_att.create_attendance(
            db,
            user_id=user_id,
            target_date=today,
            clock_in=clock_in_dt,
            clock_out=clock_out_dt,
        )

        # Create break from preset
        if break_start_dt:
            brk = crud_break.create_break(db, result.id, break_start_dt)
            if break_end_dt:
                crud_break.end_break(db, brk, break_end_dt)

        logger.info("Default-set: created attendance id=%d, preset=%s, user_id=%d", result.id, preset.name, user_id)
    db.refresh(result)
    return result


def generate_monthly_excel(db: Session, user_id: int, year: int, month: int) -> BytesIO:
    """Generate an Excel workbook for the given month's attendance records."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    records = list_attendances(db, user_id, year=year, month=month)
    # Sort ascending by date for the report
    records.sort(key=lambda r: r.date)

    wb = Workbook()
    ws = wb.active
    ws.title = "勤怠月報"

    # Title row
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"{year}年{month}月 勤怠月報"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    # Header row
    headers = ["日付", "出勤", "退勤", "休憩", "実労働時間", "入力区分", "備考"]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    def _to_dt(val):
        """Convert a value to datetime — handles both datetime objects and ISO strings."""
        if val is None:
            return None
        if isinstance(val, datetime):
            return val
        return datetime.fromisoformat(val)

    total_work_seconds = 0
    row = 4
    for rec in records:
        clock_in_val = _to_dt(rec.clock_in)
        clock_out_val = _to_dt(rec.clock_out)

        # Break total
        break_seconds = 0
        break_parts = []
        for brk in rec.breaks:
            bs = _to_dt(brk.break_start)
            be = _to_dt(brk.break_end)
            if bs and be:
                break_seconds += int((be - bs).total_seconds())
                break_parts.append(f"{bs.strftime('%H:%M')}-{be.strftime('%H:%M')}")
            elif bs:
                break_parts.append(f"{bs.strftime('%H:%M')}-")

        # Work duration
        work_str = ""
        if clock_in_val and clock_out_val:
            work_seconds = int((clock_out_val - clock_in_val).total_seconds()) - break_seconds
            if work_seconds < 0:
                work_seconds = 0
            total_work_seconds += work_seconds
            h, m = seconds_to_hm(work_seconds)
            work_str = f"{h}h {m}m"

        input_type_map = {"web": "WEB", "ic_card": "IC", "admin": "管理者"}
        input_label = input_type_map.get(rec.input_type, rec.input_type)

        ws.cell(row=row, column=1, value=str(rec.date))
        ws.cell(row=row, column=2, value=clock_in_val.strftime("%H:%M") if clock_in_val else "")
        ws.cell(row=row, column=3, value=clock_out_val.strftime("%H:%M") if clock_out_val else "")
        ws.cell(row=row, column=4, value=", ".join(break_parts) if break_parts else "")
        ws.cell(row=row, column=5, value=work_str)
        ws.cell(row=row, column=6, value=input_label)
        ws.cell(row=row, column=7, value=rec.note or "")
        row += 1

    # Total row
    total_h, total_m = seconds_to_hm(total_work_seconds)
    ws.cell(row=row, column=1, value="合計")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=5, value=f"{total_h}h {total_m}m")
    ws.cell(row=row, column=5).font = Font(bold=True)

    # Column widths
    col_widths = [12, 10, 10, 20, 14, 10, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
